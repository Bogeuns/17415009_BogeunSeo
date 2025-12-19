import os
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def find_history_csv_files():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"Scanning directory: {base_dir}")

    output_file = os.path.join(base_dir, "CL_CD_CM.xlsx")
    all_data = []
    found_files = 0

    def norm_col(c: str) -> str:
        # "       \"CMz\"      " 같은 케이스 포함해서 최대한 빡세게 정규화
        return (
            str(c)
            .replace('"', '')
            .replace("'", "")
            .replace('\t', '')
            .replace('\n', '')
            .replace('\r', '')
            .replace(' ', '')
            .strip()
            .lower()
        )

    def find_col(df, key: str):
        # key: "cl", "cd", "cmz" 같은 걸로 넣어
        # 완전일치 먼저, 없으면 포함(contains)로 백업
        norm_map = {col: norm_col(col) for col in df.columns}

        exact = [col for col, n in norm_map.items() if n == key]
        if exact:
            return exact[0]

        contains = [col for col, n in norm_map.items() if key in n]
        if contains:
            return contains[0]

        return None

    # 하위 폴더 전체 스캔 (기존 코드처럼)
    for root_dir, _, files in os.walk(base_dir):
        # 너 폴더들에서 history_direct.csv가 아니라 history.csv로 저장된 케이스가 있는 듯해서 둘 다 대응
        candidate_files = []
        if "history_direct.csv" in files:
            candidate_files.append(os.path.join(root_dir, "history_direct.csv"))
        if "history.csv" in files:
            candidate_files.append(os.path.join(root_dir, "history.csv"))

        # 둘 다 없으면, 혹시 파일명이 살짝 다른 케이스도 잡아주기 (history*.csv)
        if not candidate_files:
            for f in files:
                lf = f.lower()
                if lf.startswith("history") and lf.endswith(".csv"):
                    candidate_files.append(os.path.join(root_dir, f))

        for csv_file in candidate_files:
            try:
                df = pd.read_csv(csv_file)

                cl_col = find_col(df, "cl")
                cd_col = find_col(df, "cd")
                cmz_col = find_col(df, "cmz")

                rel_path = os.path.relpath(root_dir, base_dir)

                # AOA는 "csv가 들어있는 폴더명" 기준 (요구사항)
                aoa_name = os.path.basename(root_dir)

                # 마지막 값
                last_cl = df[cl_col].iloc[-1] if cl_col is not None and len(df) > 0 else None
                last_cd = df[cd_col].iloc[-1] if cd_col is not None and len(df) > 0 else None
                last_cmz = df[cmz_col].iloc[-1] if cmz_col is not None and len(df) > 0 else None

                # 하나라도 있으면 기록
                if any(v is not None for v in [last_cl, last_cd, last_cmz]):
                    all_data.append({
                        "AOA (Folder)": aoa_name,        # 맨 좌측열
                        "Relative Path": rel_path,
                        "CSV File": os.path.basename(csv_file),
                        "Last CL Value": last_cl,
                        "Last CD Value": last_cd,
                        "Last CMz Value": last_cmz,
                        "Total Rows": len(df),
                        "Full Path": root_dir
                    })
                    found_files += 1
                    print(f"Found in {os.path.relpath(csv_file, base_dir)}: CL={last_cl}, CD={last_cd}, CMz={last_cmz}")
                else:
                    print(f"Warning: No CL/CD/CMz columns in {os.path.relpath(csv_file, base_dir)}")
                    print(f"Available columns: {df.columns.tolist()}")

            except Exception as e:
                print(f"Error processing {os.path.relpath(csv_file, base_dir)}: {str(e)}")

    if not all_data:
        print("\nNo history*.csv files with CL/CD/CMz columns found in the directory tree.")
        return

    result_df = pd.DataFrame(all_data)

    # AOA 폴더명이 숫자면 숫자 정렬 시도
    def try_float(x):
        try:
            return float(str(x).strip())
        except:
            return None

    aoa_numeric = result_df["AOA (Folder)"].map(try_float)
    if aoa_numeric.notna().any():
        result_df["_AOA_num"] = aoa_numeric
        result_df = result_df.sort_values(by=["_AOA_num", "AOA (Folder)", "Relative Path"], na_position="last")
        result_df = result_df.drop(columns=["_AOA_num"])
    else:
        result_df = result_df.sort_values(by=["AOA (Folder)", "Relative Path"])

    # 엑셀 저장
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="CL_CD_CM", index=False)
        ws = writer.sheets["CL_CD_CM"]

        # 컬럼 너비 자동 조정
        for i, col in enumerate(result_df.columns, 1):
            max_length = max(result_df[col].astype(str).map(len).max(), len(str(col)))
            ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 60)

        # 헤더 스타일
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    print(f"\nSuccess! Found {found_files} files with CL/CD/CMz values.")
    print(f"Saved to: {output_file}")

if __name__ == "__main__":
    find_history_csv_files()
